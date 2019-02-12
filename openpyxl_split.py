# -*- coding: utf-8 -*-

from openpyxl import load_workbook
from openpyxl import Workbook

source_wb = load_workbook("input.xlsx")

source_sheet = source_wb.active

target_wb = Workbook()
target_sheet = target_wb.active

for row in source_sheet.iter_rows():
    for cell in row:
        # 内容
        target_sheet[cell.column + str(cell.row * 2 - 1)] = str(cell.value) if cell.value else ''
        # 批注
        target_sheet[cell.column + str(cell.row * 2)] = str(cell.comment.content) if cell.comment else ''

target_wb.save('split.xlsx')
