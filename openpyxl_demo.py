# -*- coding: utf-8 -*-

from openpyxl import load_workbook
from openpyxl import Workbook

source_wb = load_workbook("input.xlsx")

source_sheet = source_wb.get_sheet_by_name("Sheet1")

target_wb = Workbook()
target_sheet = target_wb.active

for row in source_sheet.iter_rows():
    for cell in row:
        target_sheet[cell.coordinate] = str(cell.comment) if cell.comment else ''

target_wb.save('output.xlsx')


# print(cell.column + '  ' + str(cell.row))
# print(cell.coordinate, cell.value)

# print(sheet)
# print(sheet["C"])
# print(sheet["4"])
# print(sheet["C4"].value)
# print(sheet.max_row)
# print(sheet.max_column)

# for row in source_sheet:
#     print('')
#     # target_sheet['A' + str(item.row)] = str(item.comment) if item.comment else ''
#     #
#     # # print(item.comment)
#     # # print('A' + str(item.row))
#     # # target_sheet['A' + str(item.row)] = item.comment
#     # # print(i.value, end=", ")
#     # # print(i.comment, end=" \r\n")
#

