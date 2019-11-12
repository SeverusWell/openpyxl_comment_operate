# -*- coding: utf-8 -*-

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.comments import Comment

source_wb = load_workbook("split.xlsx")
source_sheet = source_wb.active

target_wb = Workbook()
target_sheet = target_wb.active

target_current_row = 1

for inx, row in enumerate(source_sheet.iter_rows()):
    if 0 == inx % 2:  # 内容
        for cell in row:
            target_sheet[cell.column + str(target_current_row)] = cell.value
    else:  # 批注
        for cell in row:
            if cell.value:
                target_sheet[cell.column + str(target_current_row)].comment = Comment(cell.value, 'Python')
        target_current_row += 1

target_wb.save('source.xlsx')
